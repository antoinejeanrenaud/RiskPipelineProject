import Clean_process as cp
import var_utils as var
import Ingest as ingest
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from var_utils import calculate_parametric_var

print("Step 1: Importing Raw Data in DB")
ingest.save_raw_data_to_db()

def merge_latest_prices(positions_df: pd.DataFrame, prices_df: pd.DataFrame) -> pd.DataFrame:
    """
    For each row in positions_df, add the latest available QuoteValue
    from prices_df, matched by METAL, MATURITY, and EXCHANGE.

    Returns
    -------
    pd.DataFrame : positions_df with 'QuoteValue' and 'Price Date' columns added
    """
    # Normalize column names for consistent merge keys
    prices = prices_df.rename(columns={
        "Metal": "METAL",
        "Exchange": "EXCHANGE",
        "MTQuote": "QUOTE",
        "Price Date": "PRICE_DATE"
    }).copy()

    # Convert to datetime using exact format
    prices["PRICE_DATE"] = pd.to_datetime(prices["PRICE_DATE"], dayfirst=True, errors="coerce")


    # Drop rows with missing dates (invalid for grouping)
    prices = prices.dropna(subset=["PRICE_DATE"])

    # Get the latest price per (METAL, MATURITY, EXCHANGE)
    latest_prices = (
        prices.sort_values("PRICE_DATE")
              .groupby(["METAL", "MaturityMonth", "EXCHANGE"], as_index=False)
              .last()
    )

    # Keep only relevant columns to merge
    latest_prices = latest_prices[["METAL", "MaturityMonth", "EXCHANGE", "QUOTE", "PRICE_DATE"]]

    # Merge latest prices into positions
    merged = positions_df.merge(
        latest_prices,
        on=["METAL", "MaturityMonth", "EXCHANGE"],
        how="left"
    )

    # Rename for clarity
    merged = merged.rename(columns={
        "QUOTE": "QuoteValue",
        "PRICE_DATE": "Price Date"
    })

    return merged

def calculate_VaR_from_portfolio(positions_df, prices_df, conf=0.99, lookback = 365):
    positions_df = merge_latest_prices(positions_df, prices_df)
    # 1. Get unique combinations from positions_df
    valid_keys = positions_df[["METAL", "EXCHANGE", "MaturityMonth"]].drop_duplicates()

    # 2. Merge or filter prices_df to keep only matching rows
    filtered_prices = prices_df.rename(columns={
        "Metal": "METAL",
        "Exchange": "EXCHANGE",
    }).copy()
    filtered_prices = filtered_prices.merge(valid_keys, on=["METAL", "EXCHANGE", "MaturityMonth"], how="inner")

    filtered_prices = filtered_prices.rename(columns={
        "METAL": "Metal",
        "EXCHANGE": "Exchange",
    })

    VaR = round(var.calculate_parametric_var(positions_df, filtered_prices, conf, lookback),2)

    return VaR


def calculate_VaR(levels=["Total", "BUSINESS LINE","METAL"], conf=0.99, lookback=365, T=1):
    print("Step 2: Loading Data positions and prices from DB to perform VaR calculation.")
    positions_df, prices_df = cp.get_data()

    out = cp.detect_outliers_zscore(prices_df)
    print("###########")
    print(f"There are {out} outliers in total. ")
    print("###########")

    VaR_Total = "error"
    VaR_Levels = {}
    print("Step 3: VaR calculation...")
    for level in levels:
        if level == "Total":
            try:
                VaR_Total = calculate_VaR_from_portfolio(positions_df, prices_df, conf, lookback) * (T ** 0.5)
                print("######")
                print(f"TOTAL VAR: {VaR_Total} ")
            except Exception as e:
                print(f" Error calculating Total VaR: {e}")
        else:
            try:
                if level not in positions_df.columns:
                    print(f" Column '{level}' not found in positions data.")
                    continue

                VaR_Levels[level] = {}
                print("######")
                print(f"VaR of {level} :")
                elements = list(positions_df[level].unique())

                for el in elements:
                    sub_df = positions_df[positions_df[level] == el].copy()
                    VaR_el = calculate_VaR_from_portfolio(sub_df, prices_df, conf, lookback) * (T ** 0.5)
                    VaR_Levels[level][el] = VaR_el
                    print(f"VaR of {el} : {VaR_el} ")
                    print("######")

            except Exception as e:
                print(f" Error calculating VaR for level '{level}': {e}")




    return VaR_Total, VaR_Levels

def calculate_historical_VaR(
            levels = ["Total", "BUSINESS LINE", "METAL"],
            conf: float = 0.99,
            lookback: int = 365
    ):
        """
        Calculate Historical VaR for the full portfolio and optionally for breakdowns.

        Parameters
        ----------
        levels : list of str
            Levels in the position data to break down VaR (e.g., BUSINESS LINE, METAL).
        conf : float
            Confidence level for Historical VaR.
        lookback : int
            Lookback window in calendar days.

        Returns
        -------
        Tuple containing total VaR and a dictionary with VaR by levels.
        """
        print("Step 2: Loading data for Historical VaR...")
        positions_df, prices_df = cp.get_data()

        out = cp.detect_outliers_zscore(prices_df)
        print("###########")
        print(f"There are {out} outliers in total.")
        print("###########")

        VaR_Total = "error"
        VaR_Levels = {}

        print("Step 3: Historical VaR calculation...")
        for level in levels:
            if level == "Total":
                try:
                    VaR_Total = var.compute_historical_var(positions_df,prices_df,lookback, conf)
                    print("######")
                    print(f"TOTAL HISTORICAL VAR: {VaR_Total}")
                except Exception as e:
                    print(f"Error calculating Total Historical VaR: {e}")
            else:
                try:
                    if level not in positions_df.columns:
                        print(f"⚠️ Column '{level}' not found in positions data.")
                        continue

                    VaR_Levels[level] = {}
                    print(f"### Historical VaR by '{level}' ###")
                    elements = list(positions_df[level].unique())

                    for el in elements:
                        sub_df = positions_df[positions_df[level] == el].copy()
                        VaR_el = var.compute_historical_var(sub_df, prices_df, lookback, conf)
                        VaR_Levels[level][el] = VaR_el
                        print(f"VaR of {el}: {VaR_el}")

                except Exception as e:
                    print(f"Error calculating VaR for level '{level}': {e}")

        return VaR_Total, VaR_Levels


def run_var_analysis(historical: bool = False,
                     levels: list = ["Total", "BUSINESS LINE"],
                     conf: float = 0.99,
                     lookback: int = 365,
                     T: int = 1):
    """
    Run either historical or parametric VaR analysis.

    Parameters
    ----------
    historical : bool
        If True, compute Historical VaR; else compute Parametric VaR.
    levels : list
        List of breakdown levels (e.g., ["Total", "BUSINESS LINE", "METAL"])
    conf : float
        Confidence level (e.g., 0.99 for 99% VaR)
    lookback : int
        Lookback period in calendar days
    T : int
        Time horizon in days (used only in parametric VaR)

    Returns
    -------
    Tuple : (VaR_Total, VaR_by_Level_Dict)
    """

    if historical:
        print("Running Historical VaR...")
        return calculate_historical_VaR(levels=levels, conf=conf, lookback=lookback)
    else:
        print("Running Parametric (Normal) VaR...")
        return calculate_VaR(levels=levels, conf=conf, lookback=lookback, T=T)

# Example of how to create a class
class VaR:
        def __init__(self,
                     historical: bool = True,
                     levels: list = ["Total", "BUSINESS LINE"],
                     confidence: float = 0.99,
                     lookback_days: int = 365,
                     holding_period: int = 1):
            """
            Wrapper class for Value-at-Risk analysis (historical or parametric).

            Parameters
            ----------
            historical : bool
                If True, use historical VaR; otherwise parametric VaR.
            levels : list
                List of groupings to compute VaR by (e.g. 'BUSINESS LINE', 'METAL', etc.)
            confidence : float
                Confidence level (e.g. 0.99 for 99%)
            lookback_days : int
                Number of calendar days to look back
            holding_period : int
                Horizon in days (used only in parametric)
            """
            self.historical = historical
            self.levels = levels
            self.conf = confidence
            self.lookback = lookback_days
            self.T = holding_period

        def compute(self):
            """
            Run the VaR analysis using the configured method.

            Returns
            -------
            Tuple : (Total VaR, VaR by level dictionary)
            """
            if self.historical:
                print("Running Historical VaR...")
                return calculate_historical_VaR(
                    levels=self.levels,
                    conf=self.conf,
                    lookback=self.lookback
                )
            else:
                print("Running Parametric (Normal) VaR...")
                return calculate_VaR(
                    levels=self.levels,
                    conf=self.conf,
                    lookback=self.lookback,
                    T=self.T
                )

    # Create Excel writer
def save_var_report(var_total, var_levels, output_path="VaR_Report.xlsx"):
        """
        Save the VaR results to an Excel report with formatting and charts.
        """
        # Step 1: Save the basic DataFrames to Excel
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Total VaR
            df_total = pd.DataFrame({"VaR": [var_total]})
            df_total.to_excel(writer, sheet_name="Total_VaR", index=False)

            # VaR by level (e.g., BUSINESS LINE)
            for level_name, level_dict in var_levels.items():
                df_level = pd.DataFrame.from_dict(level_dict, orient="index", columns=["VaR"])
                df_level.index.name = level_name
                df_level.reset_index(inplace=True)
                sheet_name = f"VaR_by_{level_name.replace(' ', '_')[:31]}"
                df_level.to_excel(writer, sheet_name=sheet_name, index=False)

        # Step 2: Open workbook for formatting and charting
        wb = load_workbook(output_path)

        # Format Total_VaR
        ws_total = wb["Total_VaR"]
        ws_total["A1"] = "Total VaR (USD)"
        ws_total["A2"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        # Format and chart each VaR by level sheet
        for sheet in wb.sheetnames:
            if sheet.startswith("VaR_by_"):
                ws = wb[sheet]
                max_row = ws.max_row

                # Apply USD currency format to VaR column
                for row in range(2, max_row + 1):
                    cell = ws[f"B{row}"]
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

                # Create bar chart
                chart = BarChart()
                chart.type = "bar"
                chart.title = "VaR by " + sheet.replace("VaR_by_", "").replace("_", " ")
                chart.y_axis.title = "VaR (USD)"
                chart.x_axis.title = sheet.split("VaR_by_")[-1]

                data = Reference(ws, min_col=2, min_row=1, max_row=max_row)
                cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.width = 12
                chart.height = 6

                ws.add_chart(chart, f"D2")

        # Save enhanced Excel
        wb.save(output_path)
        print(f"✅ Enhanced VaR report saved to: {output_path}")


#positions_df, prices_df = cp.get_data()
#timeserie = var.compute_portfolio_value_time_series(positions_df, prices_df)

#var = calculate_historical_VaR()
#var.compute_historical_var(positions_df, prices_df)

result = run_var_analysis(historical=True, levels=["Total", "BUSINESS LINE", "METAL"],conf=0.99,lookback=365)

### Simple Example on how we can uses classes to compute VaR ###
#Model = VaR(historical= True,levels = ["Total", "BUSINESS LINE"],confidence = 0.99,lookback_days = 365)
#Model.compute()

print("Saving Results to Excel...")
save_var_report(result[0],result[1])

